from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List
from pathlib import Path
import uuid
import csv
import io
import asyncio
from ..database import get_db
from .. import models, schemas

router = APIRouter(
    prefix="/devices",
    tags=["devices"]
)

@router.get("/", response_model=List[schemas.DeviceOut])
def get_devices(db: Session = Depends(get_db)):
    return db.query(models.Device).all()

@router.post("/", response_model=schemas.DeviceOut, status_code=status.HTTP_201_CREATED)
def create_device(device: schemas.DeviceCreate, db: Session = Depends(get_db)):
    db_device = models.Device(**device.model_dump())
    db.add(db_device)
    db.commit()
    db.refresh(db_device)
    return db_device

@router.get("/import/template-excel")
async def get_import_template_excel(db: Session = Depends(get_db)):
    """
    Excel template dosyası döndür (kategori ve konum dropdownları ile).
    Mevcut cihazlardan kategori ve konum değerlerini çeker.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment

    # Mevcut kategorileri ve konumları al
    categories = db.query(models.Device.device_type).distinct().filter(
        models.Device.device_type.isnot(None)
    ).all()
    category_values = sorted([c[0] for c in categories if c[0]])

    locations = db.query(models.Device.location).distinct().filter(
        models.Device.location.isnot(None)
    ).all()
    location_values = sorted([l[0] for l in locations if l[0]])

    # Excel workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Cihazlar"

    # Header satırı
    headers = ["İsim", "IP Adresi", "Kategori", "Konum"]
    ws.append(headers)

    # Header formatting
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25

    # Boş satırlar ekle (örnek için 10 satır)
    for i in range(10):
        ws.append(["", "", "", ""])

    # Data validation (dropdown) oluştur
    if category_values:
        category_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(category_values)}"',
            allow_blank=False
        )
        ws.add_data_validation(category_dv)
        category_dv.add(f"C2:C1000")

    if location_values:
        location_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(location_values)}"',
            allow_blank=True
        )
        ws.add_data_validation(location_dv)
        location_dv.add(f"D2:D1000")

    # İnstructions sheet
    info_ws = wb.create_sheet("Talimatlar")
    info_ws.append(["İçe Aktarma Talimatları"])
    info_ws.append([""])
    info_ws.append(["1. İsim sütununa cihaz adı yazın (zorunlu)"])
    info_ws.append(["2. IP Adresi sütununa cihazın IP'sini yazın (zorunlu)"])
    info_ws.append(["3. Kategori ve Konum sütunlarından seçim yapın (dropdown)"])
    info_ws.append(["4. Dosyayı kaydedin ve içe aktarma sayfasında yükleyin"])

    info_title = info_ws["A1"]
    info_title.font = Font(bold=True, size=14)
    info_ws.column_dimensions["A"].width = 60

    # BytesIO'ya yaz ve döndür
    import io as io_module
    output = io_module.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cihaz_sablon.xlsx"}
    )

@router.get("/{device_id}", response_model=schemas.DeviceOut)
def get_device(device_id: int, db: Session = Depends(get_db)):
    db_device = db.query(models.Device).filter(models.Device.id == device_id).first()
    if not db_device:
        raise HTTPException(status_code=404, detail="Device not found")
    return db_device

@router.put("/{device_id}", response_model=schemas.DeviceOut)
def update_device(device_id: int, device_update: schemas.DeviceUpdate, db: Session = Depends(get_db)):
    db_device = db.query(models.Device).filter(models.Device.id == device_id).first()
    if not db_device:
        raise HTTPException(status_code=404, detail="Device not found")
    update_data = device_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_device, key, value)
    db.commit()
    db.refresh(db_device)
    return db_device

@router.delete("/{device_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_device(device_id: int, db: Session = Depends(get_db)):
    db_device = db.query(models.Device).filter(models.Device.id == device_id).first()
    if not db_device:
        raise HTTPException(status_code=404, detail="Device not found")
    # Özel ikon dosyasını sil
    if db_device.image_filename:
        img_path = Path(__file__).parents[2] / "static" / "categories" / db_device.image_filename
        try:
            if img_path.exists():
                img_path.unlink()
        except Exception:
            pass
    db.delete(db_device)
    db.commit()
    return None


@router.get("/{device_id}/image")
def get_device_image(device_id: int, db: Session = Depends(get_db)):
    """Cihaza özel ikonu serve et."""
    db_device = db.query(models.Device).filter(models.Device.id == device_id).first()
    if not db_device or not db_device.image_filename:
        raise HTTPException(status_code=404, detail="Image not found")
    file_path = Path(__file__).parents[2] / "static" / "categories" / db_device.image_filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Image file not found")
    return FileResponse(path=file_path, media_type="image/png")


MAX_IMAGE_BYTES = 2 * 1024 * 1024  # 2 MB
ALLOWED_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}
# Magic byte signatures: ext → list of accepted prefixes (None = ext'e güvenmek için extra check yok)
_IMAGE_SIGNATURES: dict[str, list[bytes]] = {
    ".png":  [b"\x89PNG\r\n\x1a\n"],
    ".jpg":  [b"\xff\xd8\xff"],
    ".jpeg": [b"\xff\xd8\xff"],
    ".gif":  [b"GIF87a", b"GIF89a"],
    ".webp": [b"RIFF"],  # RIFF....WEBP — ilk 4 byte yeterli, "WEBP" 8-12. byte
}


def _validate_image_bytes(data: bytes, ext: str) -> None:
    """Magic byte ve boyut kontrolü. SVG için XML başlığı arar."""
    if len(data) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=413, detail=f"Image too large (max {MAX_IMAGE_BYTES // 1024 // 1024} MB)")
    if ext == ".svg":
        head = data[:512].lstrip().lower()
        if not (head.startswith(b"<?xml") or head.startswith(b"<svg")):
            raise HTTPException(status_code=400, detail="Invalid SVG content")
        return
    sigs = _IMAGE_SIGNATURES.get(ext)
    if not sigs:
        raise HTTPException(status_code=400, detail=f"Unsupported image type: {ext}")
    if not any(data.startswith(sig) for sig in sigs):
        raise HTTPException(status_code=400, detail="File content does not match its extension")
    if ext == ".webp" and b"WEBP" not in data[:16]:
        raise HTTPException(status_code=400, detail="Invalid WebP content")


@router.put("/{device_id}/image")
def upload_device_image(device_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Cihaza özel ikon yükle."""
    db_device = db.query(models.Device).filter(models.Device.id == device_id).first()
    if not db_device:
        raise HTTPException(status_code=404, detail="Device not found")
    if not file.content_type or not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Only image files are allowed")

    ext = Path(file.filename or "image").suffix.lower() or ".png"
    if ext not in ALLOWED_IMAGE_EXTS:
        raise HTTPException(status_code=400, detail=f"Unsupported extension. Allowed: {', '.join(sorted(ALLOWED_IMAGE_EXTS))}")

    data = file.file.read(MAX_IMAGE_BYTES + 1)
    _validate_image_bytes(data, ext)

    # uuid4().hex tek başına path traversal'ı engeller; ext whitelist'ten geçti.
    filename = f"device_{device_id}_{uuid.uuid4().hex}{ext}"

    # Eski dosya varsa sil
    if db_device.image_filename:
        old_path = Path(__file__).parents[2] / "static" / "categories" / db_device.image_filename
        try:
            if old_path.exists():
                old_path.unlink()
        except Exception:
            pass

    file_path = Path(__file__).parents[2] / "static" / "categories" / filename
    file_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with open(file_path, "wb") as f:
            f.write(data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not save file: {str(e)}")

    db_device.image_filename = filename
    db.commit()
    db.refresh(db_device)
    return db_device


@router.delete("/{device_id}/image")
def delete_device_image(device_id: int, db: Session = Depends(get_db)):
    """Cihaza özel ikonu kaldır."""
    db_device = db.query(models.Device).filter(models.Device.id == device_id).first()
    if not db_device:
        raise HTTPException(status_code=404, detail="Device not found")
    if db_device.image_filename:
        file_path = Path(__file__).parents[2] / "static" / "categories" / db_device.image_filename
        try:
            if file_path.exists():
                file_path.unlink()
        except Exception:
            pass
        db_device.image_filename = None
        db.commit()
    return {"message": "Image removed"}

@router.post("/import/csv")
async def import_devices_csv(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    CSV dosyasından toplu cihaz ekle ve otomatik scan'ı tetikle.
    Beklenen sütunlar: name, ip_address, device_type, location (opsiyonel)
    Aynı IP'ye sahip cihazlar atlanır.
    """
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Sadece .csv dosyası kabul edilir")

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")  # BOM varsa temizle
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    # CSV text'i parse et (header'ları normalizle)
    lines = text.strip().split('\n')
    if not lines:
        raise HTTPException(status_code=400, detail="CSV dosyası boş")

    # Header satırını oku ve normalize et
    header_line = lines[0]
    raw_headers = [h.strip().strip('"') for h in header_line.split(',')]

    header_map = {
        "İsim": "name",
        "IP Adresi": "ip_address",
        "Kategori": "device_type",
        "Konum": "location",
        "name": "name",
        "ip_address": "ip_address",
        "device_type": "device_type",
        "location": "location"
    }

    normalized_headers = [header_map.get(h, h) for h in raw_headers]
    normalized_csv = ','.join(normalized_headers) + '\n' + '\n'.join(lines[1:])

    reader = csv.DictReader(io.StringIO(normalized_csv))

    added = 0
    skipped = 0
    errors = []
    added_ips = []

    for i, row in enumerate(reader, start=2):
        name = row.get("name", "").strip()
        ip   = row.get("ip_address", "").strip()
        dtype = row.get("device_type", "other").strip() or "other"
        loc  = row.get("location", "").strip()

        if not name or not ip:
            errors.append(f"Satır {i}: ad veya IP eksik")
            continue

        # IP formatı kontrolü (basit)
        parts = ip.split(".")
        if len(parts) != 4 or not all(p.isdigit() and 0 <= int(p) <= 255 for p in parts):
            errors.append(f"Satır {i}: geçersiz IP '{ip}'")
            continue

        # Duplicate IP kontrolü
        exists = db.query(models.Device).filter(models.Device.ip_address == ip).first()
        if exists:
            skipped += 1
            continue

        db_device = models.Device(
            name=name,
            ip_address=ip,
            device_type=dtype,
            location=loc or None,
        )
        db.add(db_device)
        added_ips.append(ip)
        added += 1

    db.commit()

    # Arka planda yeni cihazları scan et — ana scheduler loop'una fire-and-forget submit et,
    # böylece HTTP request handler'ı bloklanmaz ve hatalar logger'a düşer.
    if added > 0:
        from ..services.scheduler import network_scheduler
        loop = getattr(network_scheduler.scheduler, "_eventloop", None)
        if loop and not loop.is_closed():
            asyncio.run_coroutine_threadsafe(
                network_scheduler.ping_all_devices(),
                loop,
            )
        else:
            import logging
            logging.getLogger(__name__).warning(
                "Scheduler loop unavailable; imported devices will be scanned at next interval."
            )

    return {"added": added, "skipped": skipped, "errors": errors}
